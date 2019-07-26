import pandas as pd
import config as cfg
import selenium
from selenium import webdriver
import time
from bs4 import BeautifulSoup
import requests
from selenium.webdriver.common.by import By
import os
from xlwt import Workbook
import xlrd
from openpyxl import load_workbook
from selenium.common.exceptions import NoSuchElementException

index_company = 0
company = ""
ipo_prospectus = ""
file_name = cfg.company_list_filename
sheet_name = cfg.sheet_name
col_name = cfg.col_name
driver = webdriver.Chrome(executable_path= os.path.join(os.getcwd(), cfg.chrome_driver))
p_url = ""
status = ""

def clear_umwanted_data(final_list):
	try:
		final_list = list(filter(lambda x: x not in [""],final_list))
	except ValueError:
		pass
	return final_list

def save_to_excel(df):

	df = cal_details(df)

	writer = pd.ExcelWriter('Management.xlsx', engine='openpyxl')
	writer.book = load_workbook('Management.xlsx')
	if(writer.book.worksheets):
		writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
		df.to_excel(writer, startrow = writer.sheets["Sheet1"].max_row + 2, index=False)
	else:
		df.to_excel(writer, index=False)

	writer.save()

def save_financial_to_excel(df):

	writer = pd.ExcelWriter('Financial.xlsx', engine='openpyxl')
	writer.book = load_workbook('Financial.xlsx')
	if(writer.book.worksheets):

		writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
		df.to_excel(writer, startrow = writer.sheets["Sheet1"].max_row + 2, index=False, header=False)
	else:
		df.to_excel(writer, index=False, header=False)

	writer.save()

def save_total_employees():
	try :
		employees_tag = ipo_prospectus.find(lambda tag:tag.name=="b" and "Employees" in tag.text)

		if(employees_tag != None):
			employees_details = employees_tag.findNext(lambda tag:tag.name=="p" and ("full-time" in tag.text or "As of" in tag.text))
			if(employees_details == None):
				employees_details = employees_tag.findNext(lambda tag:tag.name=="div" and ("full-time" in tag.text or "As of" in tag.text))
			employees_details = employees_details.text.strip()
		else:
			employees_tag = ipo_prospectus.find(lambda tag:tag.name=="b" and "EMPLOYEES" in tag.text)
			if(employees_tag != None):
				employees_details = employees_tag.findNext(lambda tag:tag.name=="p" and ("full-time" in tag.text or "As of" in tag.text))
				if(employees_details == None):
					employees_details = employees_tag.findNext(lambda tag:tag.name=="div" and ("full-time" in tag.text or "As of" in tag.text))
				employees_details = employees_details.text.strip()
			else:
				employees_details = "Error in saving employees data!"

	except Exception as e:
		pass
	finally:
		return employees_details

def save_financial_data():
	try:
		fin_data_link =  ipo_prospectus.find(lambda tag:tag.name=="a" and "selected" in tag.text.lower())
		main_content = ipo_prospectus.find('a', attrs = {'name': fin_data_link['href'][1:]})
		table = main_content.findNext('table')
		temp_tag = table
		temp_parent = ""
		while True:
			if temp_tag == None:
				break
			elif temp_tag.name == 'p':
				if(temp_tag.find(lambda tag:tag.name=="b" and "DISCUSSION AND ANALYSIS" in tag.text.upper())  != None):
					break
				else:
					temp_parent = temp_tag
			elif temp_tag.name == 'div':
				if(temp_tag.find(lambda tag:tag.name=="b" and "DISCUSSION AND ANALYSIS" in tag.text.upper())  != None):
					break

				temp_table = temp_tag.findNext('table')
				rows = temp_table.findAll('tr')
				final_list = []
				final_list.append([company])
				for tr in rows:
					if(tr.find_all('th')):
						td = tr.find_all('th')
					else:
						td = tr.find_all('td')
					if(len(final_list) > 4):
						td = clean_col_data(td)
					row = [tr.text.strip() for tr in td]
					row = list(map(lambda b: b.replace("(",""), row))
					final_list.append(row)

				final_list = [x for x in final_list if x != []]
				final_list = pd.DataFrame(final_list)
				save_financial_to_excel(final_list)
				if(temp_table.find_parent('div') != None):
					temp_parent = temp_table.find_parent('div')
				else:
					temp_parent = temp_table
			elif temp_tag.name == 'table':
				rows = temp_tag.findAll('tr')
				final_list = []
				final_list.append([company])
				for tr in rows:
					if(tr.find_all('th')):
						td = tr.find_all('th')
					else:
						td = tr.find_all('td')
					if(len(final_list) > 4):
						td = clean_col_data(td)
					row = [tr.text.strip() for tr in td]
					row = list(map(lambda b: b.replace("(",""), row))
					final_list.append(row)

				final_list = [x for x in final_list if x != []]
				final_list = pd.DataFrame(final_list)
				save_financial_to_excel(final_list)
				if(temp_tag.find_parent('div') != None):
					temp_parent = temp_tag.find_parent('div')
				else:
					temp_parent = temp_tag

			else:
				temp_parent = temp_tag

			temp_tag = temp_parent.nextSibling

	except KeyError as e:
		pass
	except Exception as e:
		final_list = pd.DataFrame([["Error in saving financial data!"], [e]], columns=[company])
		
		save_financial_to_excel(final_list)
		pass


def clean_col_data(cols):
	try:
		cols = list(filter(lambda x: x.text.strip() not in [")", "$" , ""],cols))
	except ValueError:
		pass
	return cols

def save_management_data():
	try:
		mgm_link =  ipo_prospectus.find_all(lambda tag:tag.name=="a" and "management" in tag.text.lower())
		for i in mgm_link:
			if(i.text.strip().lower() == "management"):

				main_content = ipo_prospectus.find('a', attrs = {'name': i['href'][1:]})
				table = main_content.findNext('table')
				rows = table.findAll('tr')
				final_list = []

				for tr in rows:
					td = tr.find_all('td')
					row = [tr.text.strip() for tr in td]
					row = clear_umwanted_data(row)
					final_list.append(row)

				final_list = delete_unwanted_column(final_list)
				final_list = pd.DataFrame(final_list, columns =[company, 1, 2 ])
				final_list = add_committee_details(final_list, table)
				save_to_excel(final_list)
	except KeyError:
		pass
	except Exception as e:
		final_list = pd.DataFrame([["Error in saving Management data!"], [e]], columns=[company])
		save_to_excel(final_list)
		pass

def delete_unwanted_column(final_list):
	final_list = [x for x in final_list if x != []]
	final_list = [a[:3] for a in final_list]
	return final_list

def cal_details(df):
	try :

		total_cnt = 0
		average_age = 0

		age_list = df[1].values.tolist()
		for i in range(1,len(age_list)):
			if(age_list[i]):
				average_age+= int(age_list[i])
				total_cnt+=1

		df[2] = df[2].astype(str)

		director_count = df.loc[:,2].str.count("Director").sum()
		average_age = average_age/ total_cnt

		total = [0] * len(df)
		total[0] = total_cnt
		df['Committee Size'] = total

		board = [0] * len(df)
		board[0] = director_count
		df['Board Size'] = board

		manager = [0] * len(df)
		manager[0] = total_cnt - director_count
		df['Manager Size'] = manager

		average = [0] * len(df)
		average[0] = average_age
		df['Average Age'] = average

		employee_data = [0] * len(df)
		employee_data[0] = save_total_employees()
		df['Employee Count'] = employee_data

		return df
	except Exception as e:
		print("Error!!!!!!!!!!!!!!", company, e)
		pass
	finally:
		return df

def add_committee_details(final_list, table):
	try:
		details = [0] * len(final_list)

		for index, row  in final_list.iterrows():
			if(row[company].find('(') != -1):
				name = row[company].split('(')[0]
			elif(row[company].find('*') != -1):
				name = row[company].split('*')[0]
			else:
				name = row[company]

			employees_details = table.findNext(lambda tag:tag.name=="p" and (name in tag.text ) and ("our" in tag.text or "served as" in tag.text or "has" in tag.text))
			if(employees_details == None):
				employees_details = table.findNext(lambda tag:tag.name=="div" and (name in tag.text ) and ("our" in tag.text or "served as" in tag.text or "has" in tag.text))

			if(employees_details == None):
				new_name = name.split('\n')
				fname = new_name[0]
				lname = new_name[len(new_name)-1]
				employees_details = table.findNext(lambda tag:tag.name=="p" and (fname in tag.text  or lname in tag.text) and ("our" in tag.text or "served as" in tag.text or "has" in tag.text))
				if(employees_details == None):
					employees_details = table.findNext(lambda tag:tag.name=="div" and (fname in tag.text  or lname in tag.text ) and ("our" in tag.text or "served as" in tag.text or "has" in tag.text))

			if(employees_details == None):
				new_name = name.split(' ')
				fname = new_name[0]
				lname = new_name[len(new_name)-1]
				employees_details = table.findNext(lambda tag:tag.name=="p" and (fname in tag.text  or lname in tag.text) and ("our" in tag.text or "served as" in tag.text or "has" in tag.text))
				if(employees_details == None):
					employees_details = table.findNext(lambda tag:tag.name=="div" and (fname in tag.text  or lname in tag.text ) and ("our" in tag.text or "served as" in tag.text or "has" in tag.text))

			if(employees_details != None and name != "Name"):
				details[index] = employees_details.text.strip()
			else:
				details[index] = "Error in finding data!!!"
		final_list['details'] = details
	except :
		pass
	finally:
		return final_list

def save_ipo_prospectus():
	try :
		time.sleep(cfg.sleep_time)
		links = driver.find_elements_by_partial_link_text(cfg.text1)

		last_link = ""
		for link in links:
			last_link = link.get_attribute(cfg.link_address)

		driver.get(last_link)
		time.sleep(cfg.sleep_time)

		response = requests.get(driver.current_url)
		global p_url
		p_url = driver.current_url
		global ipo_prospectus
		ipo_prospectus = BeautifulSoup(response.content)

		save_management_data()
		save_financial_data()
		company_update('Company Found')
	except Exception as e:
		
		global status
		status += ' No IPO prospectus filed for this company!!'
		company_update(status)
		
		pass

def fetch_first_filed_prospectus():
	try :
		time.sleep(cfg.sleep_time)
		links = driver.find_elements_by_partial_link_text(cfg.text)
		last_link = ""
		for link in links:
			last_link = link.get_attribute(cfg.link_address)
		driver.get(last_link)
	except :
		
		global status
		status += ' No IPO prospectus filed for this company!!'
		company_update(status)
		
		pass

def set_company_details():
	try :
		company_name = driver.find_element_by_name(cfg.company)
		company_name.send_keys(company)
		search_button = driver.find_element_by_id(cfg.search_button_1)
		search_button.click()

	except:
		pass



def set_doc_type():
	try:
	#set document type and year
		time.sleep(cfg.sleep_time)
		type = driver.find_element_by_id(cfg.type)
		type.send_keys(cfg.doc_type)

		prior_year = driver.find_element_by_id(cfg.period)
		prior_year.send_keys(cfg.period_year)

		search_button = driver.find_element_by_xpath('//input[@type="submit"][@value="Search"]')
		search_button.click()

	except:
		
		global status
		status += ' Company Not Found'
		company_update(status)
		
		pass

def company_update(status):
	print(p_url,"----", company)
	df = pd.DataFrame([status])
	df1 = pd.DataFrame([p_url])
	writer = pd.ExcelWriter(file_name, engine='openpyxl')
	writer.book = load_workbook(file_name)
	writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
	if(p_url != None):
		df1.to_excel(writer, sheet_name=sheet_name ,startrow=index_company+1, startcol=3, index=False, header=False)
	df.to_excel(writer, sheet_name=sheet_name ,startrow=index_company+1, startcol=2, index=False, header=False)
	writer.save()


def read_company_list() :
	df = pd.read_excel(file_name, sheet_name=sheet_name)
	return df

def main():
	dataframe = read_company_list()
	
	for index, row  in dataframe.iterrows():
	
		if(pd.isnull(row["Status"])):
		
			global company
			company = row["Company Name"]
			
			global index_company
			index_company = index
			driver.get(cfg.web_url)
			set_company_details()
			set_doc_type()
			fetch_first_filed_prospectus()
			save_ipo_prospectus()
			time.sleep(5)
			driver.close


main()
